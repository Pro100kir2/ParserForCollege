import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import List
from src.domain.interfaces import Exporter
from src.domain.models import ParseResult, ExportConfig, SectionType
from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelExporter(Exporter):
    def export(self, result: ParseResult, config: ExportConfig) -> List[str]:
        os.makedirs(config.output_dir, exist_ok=True)
        output_files = []
        
        if config.single_workbook:
            output_file = self._export_single_workbook(result, config)
            output_files.append(output_file)
        else:
            output_files.extend(self._export_multiple_files(result, config))
        
        logger.info(f"Exported {len(output_files)} file(s) to {config.output_dir}")
        return output_files
    
    def _export_single_workbook(self, result: ParseResult, config: ExportConfig) -> str:
        wb = Workbook()
        wb.remove(wb.active)
        
        self._add_sheet(wb, "Main Literature", result.main_literature)
        self._add_sheet(wb, "Additional Literature", result.additional_literature)
        self._add_sheet(wb, "Material Resources", result.material_resources)
        
        output_path = os.path.join(config.output_dir, config.combined_filename)
        wb.save(output_path)
        logger.info(f"Created single workbook: {output_path}")
        return output_path
    
    def _export_multiple_files(self, result: ParseResult, config: ExportConfig) -> List[str]:
        output_files = []
        
        main_path = os.path.join(config.output_dir, config.filename_main)
        self._export_single_sheet(main_path, result.main_literature, "Main Literature")
        output_files.append(main_path)
        
        additional_path = os.path.join(config.output_dir, config.filename_additional)
        self._export_single_sheet(additional_path, result.additional_literature, "Additional Literature")
        output_files.append(additional_path)
        
        material_path = os.path.join(config.output_dir, config.filename_material)
        self._export_single_sheet(material_path, result.material_resources, "Material Resources")
        output_files.append(material_path)
        
        return output_files
    
    def _export_single_sheet(self, file_path: str, items: List[str], sheet_name: str):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        self._populate_sheet(ws, items)
        wb.save(file_path)
        logger.info(f"Created sheet: {file_path}")
    
    def _add_sheet(self, workbook: Workbook, sheet_name: str, items: List[str]):
        ws = workbook.create_sheet(title=sheet_name)
        self._populate_sheet(ws, items)
    
    def _populate_sheet(self, worksheet, items: List[str]):
        # Header
        header = worksheet.cell(row=1, column=1, value="Entry")
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal='left')
        
        # Data
        for idx, item in enumerate(items, start=2):
            cell = worksheet.cell(row=idx, column=1, value=item)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column width
        worksheet.column_dimensions['A'].width = 80
